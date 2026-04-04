from __future__ import annotations
"""
Semi-Intel API Routes
======================
대시보드에 필요한 모든 엔드포인트
"""

from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Request, Query, HTTPException
from loguru import logger

router = APIRouter()


def get_db(request: Request):
    return request.app.state.db


# ============================================================
# Health & Status
# ============================================================

@router.get("/status")
def get_status(request: Request):
    db = get_db(request)
    stats = db.get_collection_stats()
    
    pipeline_state = {}
    if hasattr(request.app.state, "pipeline"):
        pipe = request.app.state.pipeline
        pipeline_state = {
            "is_running": pipe.is_running,
            "last_run_time": str(pipe.last_run_time) if pipe.last_run_time else None,
            "next_run_time": str(pipe.next_run_time) if pipe.next_run_time else None,
        }

    return {
        "status": "ok",
        "total_records": stats["total_records"],
        "series_count": stats["series_count"],
        "latest_collection": str(stats["latest_collection"]) if stats["latest_collection"] else None,
        "pipeline": pipeline_state
    }


# ============================================================
# Indicators
# ============================================================

@router.get("/indicators")
def list_indicators(request: Request):
    """전체 지표 메타데이터 목록"""
    from config.indicators import ALL_INDICATORS
    return [
        {
            "id": ind.id,
            "name": ind.name,
            "tier": str(ind.tier.value),
            "category": ind.category,
            "source": ind.source,
            "frequency": ind.frequency.value,
            "dimension": ind.dimension.value,
            "book_chapter": ind.book_chapter,
            "semi_relevance": ind.semi_relevance,
            "signal_logic": ind.signal_logic,
            "fred_series": [{"code": s.code, "name": s.name} for s in ind.fred_series],
            "yahoo_symbols": ind.yahoo_symbols,
            "lag_days": ind.lag_days,
        }
        for ind in ALL_INDICATORS
    ]


@router.get("/indicators/{indicator_id}")
def get_indicator(indicator_id: str, request: Request):
    """개별 지표 상세 + 최근 데이터"""
    from config.indicators import get_indicator as find_ind
    ind = find_ind(indicator_id)
    if not ind:
        raise HTTPException(404, f"Indicator not found: {indicator_id}")

    db = get_db(request)

    # 각 시리즈의 최근 데이터
    series_data = {}
    for s in ind.fred_series:
        data = db.get_series_data(s.code)
        if data:
            recent = data[-60:]  # 최근 60개
            series_data[s.code] = {
                "name": s.name,
                "latest_value": recent[-1]["value"],
                "latest_date": str(recent[-1]["date"]),
                "data_points": len(data),
                "recent": [{"date": str(d["date"]), "value": d["value"]} for d in recent],
            }

    for sym in ind.yahoo_symbols:
        data = db.get_series_data(sym)
        if data:
            recent = data[-60:]
            series_data[sym] = {
                "name": sym,
                "latest_value": recent[-1]["value"],
                "latest_date": str(recent[-1]["date"]),
                "data_points": len(data),
                "recent": [{"date": str(d["date"]), "value": d["value"]} for d in recent],
            }

    return {
        "id": ind.id,
        "name": ind.name,
        "tier": str(ind.tier.value),
        "category": ind.category,
        "dimension": ind.dimension.value,
        "semi_relevance": ind.semi_relevance,
        "signal_logic": ind.signal_logic,
        "series_data": series_data,
    }


# ============================================================
# Time Series Data
# ============================================================

@router.get("/series/{series_code}")
def get_series(series_code: str, request: Request,
               start: Optional[str] = None,
               end: Optional[str] = None,
               limit: int = Query(default=500, le=5000)):
    """시계열 데이터 조회"""
    db = get_db(request)

    start_date = datetime.strptime(start, "%Y-%m-%d").date() if start else None
    end_date = datetime.strptime(end, "%Y-%m-%d").date() if end else None

    data = db.get_series_data(series_code, start_date=start_date, end_date=end_date)
    if not data:
        raise HTTPException(404, f"No data for series: {series_code}")

    data = data[-limit:]
    return {
        "series_code": series_code,
        "count": len(data),
        "data": [{"date": str(d["date"]), "value": d["value"]} for d in data],
    }


# ============================================================
# Signals
# ============================================================

@router.get("/signals")
def generate_signals(request: Request):
    """전체 시그널 생성"""
    db = get_db(request)
    from analysis.signal_generator import SignalGenerator

    gen = SignalGenerator(db)
    signals = gen.generate_all()

    result = {
        "generated_at": datetime.utcnow().isoformat(),
        "count": len(signals),
        "signals": {},
        "summary": {"bullish": 0, "bearish": 0, "neutral": 0},
    }

    for ind_id, sig in signals.items():
        result["signals"][ind_id] = {
            "signal_type": sig.signal_type,
            "strength": round(sig.strength, 3),
            "dimension": sig.dimension,
            "sub_signals": sig.sub_signals,
            "description": sig.description,
        }
        result["summary"][sig.signal_type] += 1

    return result


@router.get("/signals/{indicator_id}")
def generate_signal(indicator_id: str, request: Request):
    """단일 지표 시그널"""
    db = get_db(request)
    from analysis.signal_generator import SignalGenerator

    gen = SignalGenerator(db)
    sig = gen.generate(indicator_id)
    if not sig:
        raise HTTPException(404, f"No signal for: {indicator_id}")

    return {
        "indicator_id": sig.indicator_id,
        "signal_type": sig.signal_type,
        "strength": round(sig.strength, 3),
        "dimension": sig.dimension,
        "sub_signals": sig.sub_signals,
        "description": sig.description,
    }


# ============================================================
# Composite Score
# ============================================================

@router.get("/score")
def get_composite_score(request: Request):
    """Semiconductor Cycle Composite Score 산출"""
    db = get_db(request)
    from analysis.composite_score import CompositeScoreCalculator

    calc = CompositeScoreCalculator(db)
    result = calc.calculate()
    calc.save_to_db(result)

    return {
        "date": str(result.date),
        "total_score": result.total_score,
        "regime": result.regime,
        "regime_description": result.regime_description,
        "investment_action": result.investment_action,
        "trend_alert": result.trend_alert,
        "confidence_level": result.confidence_level,
        "data_coverage": result.data_coverage,
        "signal_count": result.signal_count,
        "dimensions": {
            name: {
                "score": dim.score,
                "weight": dim.weight,
                "confidence": dim.confidence,
                "contributing_signals": dim.contributing_signals,
            }
            for name, dim in result.dimensions.items()
        },
    }


@router.get("/score/history")
def get_score_history(request: Request, days: int = Query(default=90, le=365)):
    """Composite Score 히스토리"""
    db = get_db(request)
    from db.database import CompositeScore
    session = db.get_session()
    try:
        cutoff = date.today() - timedelta(days=days)
        records = (session.query(CompositeScore)
                   .filter(CompositeScore.date >= cutoff)
                   .order_by(CompositeScore.date)
                   .all())
        return {
            "count": len(records),
            "history": [
                {
                    "date": str(r.date),
                    "total_score": r.total_score,
                    "demand_score": r.demand_score,
                    "supply_score": r.supply_score,
                    "price_score": r.price_score,
                    "macro_score": r.macro_score,
                    "global_score": r.global_score,
                    "regime": r.regime,
                }
                for r in records
            ],
        }
    finally:
        session.close()


# ============================================================
# Scenarios
# ============================================================

@router.get("/scenarios")
def list_scenarios(request: Request):
    """사용 가능한 시나리오 목록"""
    db = get_db(request)
    from analysis.scenario_analyzer import ScenarioAnalyzer
    analyzer = ScenarioAnalyzer(db)
    return {"scenarios": analyzer.list_scenarios()}


@router.get("/scenarios/compare")
def compare_scenarios(request: Request):
    """전체 시나리오 비교 분석"""
    db = get_db(request)
    from analysis.scenario_analyzer import ScenarioAnalyzer
    from analysis.composite_score import CompositeScoreCalculator

    calc = CompositeScoreCalculator(db)
    base_result = calc.calculate()

    analyzer = ScenarioAnalyzer(db)
    comparison = analyzer.compare_scenarios(base_result=base_result)

    return {
        "base_score": comparison.base_score,
        "scenarios": comparison.scenarios,
    }


@router.get("/scenarios/{scenario_id}")
def analyze_scenario(scenario_id: str, request: Request):
    """개별 시나리오 상세 분석"""
    db = get_db(request)
    from analysis.scenario_analyzer import ScenarioAnalyzer

    analyzer = ScenarioAnalyzer(db)
    result = analyzer.analyze_scenario(scenario_id)
    if not result:
        raise HTTPException(404, f"Unknown scenario: {scenario_id}")
    return result


# ============================================================
# Briefing
# ============================================================

@router.get("/briefing")
def get_briefing(request: Request):
    """전체 투자 브리핑 (JSON)"""
    db = get_db(request)
    from analysis.briefing import BriefingGenerator
    gen = BriefingGenerator(db)
    return gen.generate_full_briefing()


# ============================================================
# Data Collection Trigger
# ============================================================

@router.post("/collect/fred")
def trigger_fred_collection(request: Request):
    """FRED 데이터 수집 트리거"""
    db = get_db(request)
    fred_key = os.getenv("FRED_API_KEY", "")
    if not fred_key or fred_key == "your_fred_api_key_here":
        raise HTTPException(400, "FRED_API_KEY not configured")

    from collectors.fred_collector import FredCollector
    collector = FredCollector(fred_key, db)
    results = collector.collect_all()

    total = sum(sum(r["added"] for r in res["series_results"]) for res in results)
    return {"status": "completed", "indicators": len(results), "new_records": total}


@router.post("/collect/yahoo")
def trigger_yahoo_collection(request: Request):
    """Yahoo Finance 수집 트리거"""
    db = get_db(request)
    from collectors.yahoo_collector import YahooCollector
    collector = YahooCollector(db)
    results = collector.collect_all()
    return {"status": "completed", "indicators": len(results)}


# ============================================================
# AI Advisory (Phase 4)
# ============================================================

def _get_engine(db):
    """.env 설정 기반으로 AdvisoryEngine 생성 (google/openai/anthropic)"""
    import os
    from fastapi import HTTPException
    from advisory.engine import AdvisoryEngine
    
    provider = os.getenv("LLM_PROVIDER", "google").lower()
    key_map = {
        "anthropic": "ANTHROPIC_API_KEY",
        "openai": "OPENAI_API_KEY",
        "google": "GOOGLE_API_KEY",
    }
    env_key = key_map.get(provider, "GOOGLE_API_KEY")
    api_key = os.getenv(env_key, "")
    
    if not api_key:
        raise HTTPException(status_code=500, detail=f"{env_key} not configured in .env (Provider: {provider})")
        
    return AdvisoryEngine(db, provider=provider, api_key=api_key)


@router.post("/ai/ask")
def ai_ask(request: Request, body: dict):
    """
    AI 자유 질의

    Body: {"question": "HBM 수요 전망은?", "conversation_id": "optional"}
    """
    db = get_db(request)
    question = body.get("question", "")
    if not question:
        raise HTTPException(400, "question field required")

    engine = _get_engine(db)
    response = engine.ask(question)

    return {
        "question": question,
        "response": response,
        "usage": engine.get_usage(),
    }


@router.get("/ai/briefing")
def ai_briefing(request: Request):
    """AI 일일 브리핑"""
    db = get_db(request)
    engine = _get_engine(db)
    content = engine.daily_briefing()

    return {
        "date": str(date.today()),
        "briefing": content,
        "usage": engine.get_usage(),
    }


@router.get("/ai/indicator/{indicator_id}")
def ai_indicator_analysis(indicator_id: str, request: Request):
    """AI 지표 심층 분석"""
    db = get_db(request)
    engine = _get_engine(db)
    analysis = engine.analyze_indicator(indicator_id)

    return {
        "indicator_id": indicator_id,
        "analysis": analysis,
        "usage": engine.get_usage(),
    }


@router.get("/ai/scenario/{scenario_id}")
def ai_scenario_analysis(scenario_id: str, request: Request):
    """AI 시나리오 심층 분석"""
    db = get_db(request)
    engine = _get_engine(db)
    analysis = engine.analyze_scenario(scenario_id)

    return {
        "scenario_id": scenario_id,
        "analysis": analysis,
        "usage": engine.get_usage(),
    }


@router.get("/ai/regime")
def ai_regime_analysis(request: Request):
    """AI Regime 전환 분석"""
    db = get_db(request)
    engine = _get_engine(db)
    analysis = engine.analyze_regime_transition()

    return {
        "analysis": analysis,
        "usage": engine.get_usage(),
    }


# ============================================================
# Export (CSV / XLSX)
# ============================================================

@router.get("/export/score.csv")
def export_score_csv(request: Request):
    """Composite Score 히스토리 CSV 내보내기"""
    import csv, io
    db = get_db(request)
    from db.database import CompositeScore
    session = db.get_session()
    try:
        records = session.query(CompositeScore).order_by(CompositeScore.date).all()
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["date", "total_score", "demand", "supply", "price", "macro", "global", "regime"])
        for r in records:
            writer.writerow([r.date, r.total_score, r.demand_score, r.supply_score,
                             r.price_score, r.macro_score, r.global_score, r.regime])
        from fastapi.responses import StreamingResponse
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=semi_intel_scores.csv"}
        )
    finally:
        session.close()


@router.get("/export/signals.csv")
def export_signals_csv(request: Request):
    """현재 시그널 CSV 내보내기"""
    import csv, io
    db = get_db(request)
    from analysis.signal_generator import SignalGenerator
    gen = SignalGenerator(db)
    signals = gen.generate_all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["indicator_id", "signal_type", "strength", "dimension", "description"])
    for ind_id, sig in signals.items():
        writer.writerow([ind_id, sig.signal_type, round(sig.strength, 3), sig.dimension, sig.description])

    from fastapi.responses import StreamingResponse
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=semi_intel_signals.csv"}
    )


@router.get("/export/indicators.csv")
def export_indicators_csv(request: Request):
    """지표 메타데이터 + 최신 값 CSV"""
    import csv, io
    db = get_db(request)
    from config.indicators import ALL_INDICATORS

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["indicator_id", "name", "tier", "category", "source", "frequency",
                     "dimension", "series_code", "latest_value", "latest_date"])
    for ind in ALL_INDICATORS:
        series_list = [(s.code, s.name) for s in ind.fred_series] + [(s, s) for s in ind.yahoo_symbols]
        if not series_list:
            writer.writerow([ind.id, ind.name, str(ind.tier.value), ind.category, ind.source,
                             ind.frequency.value, ind.dimension.value, "", "", ""])
        for code, _ in series_list:
            data = db.get_series_data(code)
            if data:
                latest = data[-1]
                writer.writerow([ind.id, ind.name, str(ind.tier.value), ind.category, ind.source,
                                 ind.frequency.value, ind.dimension.value, code,
                                 latest["value"], latest["date"]])
            else:
                writer.writerow([ind.id, ind.name, str(ind.tier.value), ind.category, ind.source,
                                 ind.frequency.value, ind.dimension.value, code, "", ""])

    from fastapi.responses import StreamingResponse
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=semi_intel_indicators.csv"}
    )


@router.get("/export/full.xlsx")
def export_full_xlsx(request: Request):
    """전체 데이터 XLSX 내보내기 (3개 시트: Score History, Signals, Indicators)"""
    import io
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl 필요. 실행: pip install openpyxl")

    db = get_db(request)

    wb = openpyxl.Workbook()

    # Sheet 1: Score History
    ws1 = wb.active
    ws1.title = "Score History"
    ws1.append(["Date", "Total Score", "Demand", "Supply", "Price", "Macro", "Global", "Regime"])
    from db.database import CompositeScore
    session = db.get_session()
    try:
        records = session.query(CompositeScore).order_by(CompositeScore.date).all()
        for r in records:
            ws1.append([str(r.date), r.total_score, r.demand_score, r.supply_score,
                        r.price_score, r.macro_score, r.global_score, r.regime])
    finally:
        session.close()

    # Sheet 2: Current Signals
    ws2 = wb.create_sheet("Signals")
    ws2.append(["Indicator ID", "Signal Type", "Strength", "Dimension", "Description"])
    from analysis.signal_generator import SignalGenerator
    gen = SignalGenerator(db)
    signals = gen.generate_all()
    for ind_id, sig in signals.items():
        ws2.append([ind_id, sig.signal_type, round(sig.strength, 3), sig.dimension, sig.description])

    # Sheet 3: Indicators + Latest Values
    ws3 = wb.create_sheet("Indicators")
    ws3.append(["Indicator ID", "Name", "Tier", "Category", "Dimension", "Series Code", "Latest Value", "Latest Date"])
    from config.indicators import ALL_INDICATORS
    for ind in ALL_INDICATORS:
        series_list = [(s.code, s.name) for s in ind.fred_series] + [(s, s) for s in ind.yahoo_symbols]
        for code, _ in series_list:
            data = db.get_series_data(code)
            if data:
                latest = data[-1]
                ws3.append([ind.id, ind.name, str(ind.tier.value), ind.category,
                            ind.dimension.value, code, latest["value"], str(latest["date"])])

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=semi_intel_export.xlsx"}
    )


import os
from datetime import date
